import pandas as pd
import re
from openpyxl import load_workbook


def classify_data(entry):
    # 提取品牌和门店信息
    brand_match = re.search(r'(.+?)（(.+?)）', entry)
    if brand_match and not re.search(r'\d+\.\d+', brand_match.group(2)):
        brand = brand_match.group(1)
        store = brand_match.group(2)
    else:
        brand = re.match(r'(.+?)(?=上海市|星级评分|访问人数|营收|热门榜)', entry).group(0).strip()
        store = ''

    # 提取案例种类
    if '热门榜' in entry:
        item_type = '热门榜'
    elif '星级评分' in entry:
        item_type = '星级'
    elif '访问人数' in entry:
        item_type = '访问人数'
    elif '营收' in entry:
        item_type = '营收'
    else:
        item_type = '未知'

    # 提取案例详情
    if brand_match and not re.search(r'\d+\.\d+', brand_match.group(2)):
        details = entry.split('）', 1)[1].split('（', 1)[0].strip()
    else:
        details = entry.split(item_type, 1)[1].strip()
        if '（' in details:
            details = details.split('（')[0].strip()

    return brand, store, item_type, details


def main(input_file_path, sheet_name, output_file_path):
    # 读取Excel文件
    operational_cases = pd.read_excel(input_file_path, sheet_name=sheet_name, header=None)

    # 处理数据
    processed_data = [classify_data(entry[0]) for entry in operational_cases.values]

    # 创建 DataFrame
    df = pd.DataFrame(processed_data, columns=['品牌', '门店', '案例种类', '案例详情'])

    # 使用openpyxl加载工作簿
    wb = load_workbook(input_file_path)
    ws = wb['陈雪']

    # 将解析后的数据写入工作表的A21开始的格子
    for i, row in df.iterrows():
        ws.cell(row=21 + i, column=1, value=row['品牌'])
        ws.cell(row=21 + i, column=2, value=row['门店'])
        ws.cell(row=21 + i, column=3, value=row['案例种类'])
        ws.cell(row=21 + i, column=4, value=row['案例详情'])

    # 保存更新后的工作簿
    wb.save(output_file_path)

    print(f"文件已经保存为 {output_file_path}")


if __name__ == "__main__":
    input_file_path = 'C:/Users/豆豆/Desktop/perfomance/perfomancesource/7月绩效考核表-陈雪.xlsx'  # 输入文件路径
    sheet_name = '运营案例'  # 数据工作表名称
    output_file_path = 'C:/Users/豆豆/Desktop/perfomance/perfomanceoutput/7月绩效考核-陈雪update.xlsx'  # 输出文件路径
    main(input_file_path, sheet_name, output_file_path)
