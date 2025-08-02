"""
日报生成主入口脚本
功能：从数据库拉取最新数据，生成日报并输出到指定目录。
所有路径、品牌映射等配置均引用 config。
"""
import os
import json
import pandas as pd
import time
from datetime import datetime, timedelta
from pathlib import Path
from sqlalchemy import create_engine
import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.config import db_config, REPORT_OUTPUT_DIR
from utils.logger import get_logger

# 数据库连接
DB_CONNECTION_STRING = f'mysql+pymysql://{db_config["user"]}:{db_config["password"]}@{db_config["host"]}:{db_config["port"]}/{db_config["database"]}'
engine = create_engine(DB_CONNECTION_STRING)

# 配置参数
THRESHOLD = 30   # 暴涨/暴跌判定阈值 %
CORE_FIELDS = ["消费金额", "打卡人数", "新增收藏人数", "新好评数", "新中差评数"]
RANK_PLATFORMS = {
    "dianping_hot": "点评热门榜",
    "dianping_checkin": "点评打卡人气榜",
    "dianping_rating":"点评好评榜"
}

def get_store_mapping():
    """获取门店映射数据"""
    try:
        store_map = pd.read_sql(
            "SELECT 门店ID AS store_id, 美团门店ID, 推广门店 AS brand_name, 运营师 AS operator FROM store_mapping",
            engine
        )
        return store_map
    except Exception as e:
        logger = get_logger('generate_daily_report')
        logger.error(f"获取门店映射失败: {e}")
        return pd.DataFrame()

def summarize(df, fields):
    """
    根据字段汇总数据，如果字段缺失则跳过
    """
    summary = {}
    for col in fields:
        if col not in df.columns:
            continue
        # 数值型字段求和，其他字段取平均值
        if any(x in col for x in ['金额', '人数', '次数', '笔数', '数']):
            value = df[col].sum()
        else:
            value = df[col].mean()
        summary[col] = round(value, 2) if isinstance(value, float) else value
    return summary

def compute_cpc_contribution_ratios(op_summary, cpc_summary):
    """计算CPC贡献比例"""
    def safe_div(n, d, precision=2):
        return round(n / d * 100, precision) if d else None

    clicks = cpc_summary.get("clicks", 0)
    impressions = cpc_summary.get("impressions", 0)
    cost = cpc_summary.get("cost", 0)
    orders = cpc_summary.get("orders", 0)

    ratios = {
        "曝光占比": safe_div(impressions, op_summary.get("曝光人数", 0)),
        "点击占比": safe_div(clicks, op_summary.get("访问人数", 0)),
        "平均点击单价": round(cost / clicks, 4) if clicks else 0,
        "CPA": round(cost / orders, 2) if orders else None,
        "CTR": safe_div(clicks, impressions),
        "CVR": safe_div(orders, clicks)
    }
    return ratios

def is_big_change(pct_str, threshold=THRESHOLD):
    """判断是否为大幅变化"""
    try:
        return abs(float(pct_str.strip('%'))) >= threshold
    except:
        return False

def fmt(val, pct_str):
    """格式化数值和百分比"""
    if pct_str == "N/A":
        return str(val)
    pct = float(pct_str.strip('%'))
    if pct > 0:
        return f"{val}(+{pct}%)"
    if pct < 0:
        return f"{val}(-{abs(pct)}%)"
    return f"{val}(持平)"

# ==================== 图表生成功能（已注释） ====================
# TODO: 用户要求暂时忽略图片功能，后续会大改
# 如需重新启用，请取消以下注释并删除 pass 语句
def plot_vertical_table(brand, report_date, engine):
    """生成品牌最近7天关键指标图表（已禁用）"""
    pass
    # 原始代码已注释，重新启用时请取消注释以下代码块
    """
    logger = get_logger('generate_daily_report')
    try:
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        
        # 注册中文字体
        try:
            font_manager.fontManager.addfont("./fonts/SimHei.ttf")
            plt.rcParams['font.family'] = 'SimHei'
        except:
            plt.rcParams['font.family'] = ['SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 取最近7天数据
        start = report_date - timedelta(days=6)
        df = pd.read_sql(f'''
            SELECT 
                `日期`,
                `消费金额` AS `消费金额`,
                `曝光人数`,
                `访问人数`,
                `购买人数`,
                `打卡人数`,
                `扫码人数`,
                `新增收藏人数`,
                `新好评数`,
                `新中差评数`,
                `点评星级` AS `星级`
            FROM `operation_data`
            WHERE `美团门店ID` IN (
                SELECT `美团门店ID` 
                  FROM `store_mapping` 
                 WHERE `推广门店` = '{brand}'
            )
              AND `日期` BETWEEN '{start.date()}' AND '{report_date.date()}'
            ORDER BY `日期`
        ''', engine)
        
        if df.empty:
            logger.warning(f"⚠️ {brand} 最近7天无数据，跳过图表生成")
            return
            
        # 把日期列转成纯 YYYY-MM-DD 的字符串
        df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y-%m-%d')
        
        # 衍生星期几
        weekday_map = {0:'星期一',1:'星期二',2:'星期三',3:'星期四',
                       4:'星期五',5:'星期六',6:'星期日'}
        df['星期'] = pd.to_datetime(df['日期']).dt.weekday.map(weekday_map)
        
        # 强制把所有"人数"列和评分列变成整数
        int_cols = [
            '曝光人数','访问人数','购买人数',
            '打卡人数','扫码人数','新增收藏人数',
            '新好评数','新中差评数'
        ]
        df[int_cols] = df[int_cols].astype(int)
        
        # 计算两段转化率
        df['曝光-访问转化率'] = (df['访问人数'] / df['曝光人数'] * 100).round(1).astype(str) + '%'
        df['访问-购买转化率'] = (df['购买人数'] / df['访问人数'] * 100).round(1).astype(str) + '%'
        
        # 最终选列 & 排序，并重命名为短标题
        cols = [
            '日期', '星期', '消费金额', '曝光人数', '访问人数', '购买人数',
            '曝光-访问转化率', '访问-购买转化率',
            '新增收藏人数', '打卡人数',
            '新好评数', '新中差评数', '扫码人数', '星级'
        ]
        df = df[cols].rename(columns={
            '消费金额': '消费',
            '曝光人数': '曝光',
            '访问人数': '访问',
            '购买人数': '购买',
            '曝光-访问转化率': '访问转化',
            '访问-购买转化率': '购买转化',
            '新增收藏人数': '收藏',
            '打卡人数': '打卡',
            '新好评数': '好评',
            '新中差评数': '差评',
            '扫码人数': '扫码',
            '星级':'星级'
        })
        
        # 绘制表格
        col_widths = []
        for c in df.columns:
            if c in ['日期', '星期']:
                col_widths.append(0.08)
            elif c in ['消费', '曝光', '访问', '购买', '访问转化', '购买转化']:
                col_widths.append(0.06)
            else:
                col_widths.append(0.05)
        
        fig, ax = plt.subplots(figsize=(16, 0.6 * len(df) + 1))
        ax.axis('off')
        tbl = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            colWidths=col_widths,
            cellLoc='center',
            loc='center'
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(11)
        tbl.scale(1, 1.2)
        
        ax.set_title(f"{brand} 最近7天关键指标", fontsize=16, pad=12)
        plt.tight_layout()
        
        # 确保输出目录存在
        out_dir = Path(REPORT_OUTPUT_DIR)
        out_dir.mkdir(exist_ok=True)
        out = out_dir / f"{brand}_最近7天指标表.png"
        fig.savefig(out, dpi=150, bbox_inches='tight')
        plt.close(fig)
        logger.info(f"✅ 保存图表：{out}")
        
    except ImportError:
        logger.warning("matplotlib未安装，跳过图表生成")
    except Exception as e:
        logger.error(f"生成图表失败: {e}")
    """

def analyze_brand_performance(brand, df_op, op_sum, link_ratio):
    """分析品牌表现并给出价值洞察"""
    insights = []
    
    # 消费金额分析
    rev_val = round(df_op.iloc[0]['消费金额'], 0)
    rev_pct = link_ratio.get('消费金额', 'N/A')
    if rev_pct != 'N/A':
        rev_change = float(rev_pct.strip('%'))
        if rev_change > 20:
            insights.append("💰 消费金额大幅增长，表现优秀")
        elif rev_change < -20:
            insights.append("⚠️ 消费金额下降明显，需要关注")
    
    # 转化率分析
    exposure = df_op.iloc[0]['曝光人数']
    visits = df_op.iloc[0]['访问人数']
    purchases = df_op.iloc[0]['购买人数']
    
    if exposure > 0 and visits > 0:
        visit_rate = visits / exposure
        if visit_rate < 0.05:
            insights.append("📉 访问转化率偏低，建议优化页面")
        elif visit_rate > 0.15:
            insights.append("📈 访问转化率优秀")
    
    if visits > 0 and purchases > 0:
        purchase_rate = purchases / visits
        if purchase_rate < 0.02:
            insights.append("🛒 购买转化率偏低，建议优化商品展示")
        elif purchase_rate > 0.08:
            insights.append("🎯 购买转化率优秀")
    
    # 评价分析
    good_val = int(df_op.iloc[0]['新好评数'])
    bad_val = int(df_op.iloc[0]['新中差评数'])
    
    if good_val > 0 and bad_val == 0:
        insights.append("⭐ 好评稳定，无差评，服务表现优秀")
    elif bad_val > 0:
        insights.append("⚠️ 出现差评，需要及时跟进处理")
    
    # 收藏分析
    col_val = int(op_sum.get("新增收藏人数", 0))
    col_pct = link_ratio.get("新增收藏人数", "N/A")
    if col_pct != 'N/A':
        col_change = float(col_pct.strip('%'))
        if col_change > 50:
            insights.append("❤️ 收藏增长显著，用户粘性提升")
        elif col_change < -50:
            insights.append("📉 收藏下降明显，需要关注用户留存")
    
    return insights

# ==================== 对比表生成功能（已注释） ====================
# TODO: 用户要求暂时忽略图片功能，后续会大改
# 如需重新启用，请取消以下注释并删除 pass 语句
def plot_comparison_table(brand, report_date):
    """生成品牌同期对比表格（已禁用）"""
    pass
    # 原始代码已注释，重新启用时请取消注释以下代码块
    """
    logger = get_logger('generate_daily_report')
    try:
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        
        # 注册中文字体
        try:
            font_manager.fontManager.addfont("./fonts/SimHei.ttf")
            plt.rcParams['font.family'] = 'SimHei'
        except:
            plt.rcParams['font.family'] = ['SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        # 从store_map取出本品牌所有美团门店ID
        mids = store_map.loc[store_map['brand_name'] == brand, '美团门店ID'].astype(str).tolist()
        if not mids:
            logger.warning(f"⚠️ 品牌 {brand} 无对应美团门店ID，跳过对比表")
            return
        mids_sql = ",".join(f"'{m}'" for m in mids)
        
        # 计算"本期""上期"日期
        wd = report_date.weekday()
        if wd == 0:
            # 周一：本期 = 周五~周日， 上期 = 上周周五~周日
            curr_start = report_date - timedelta(days=3)
            prev_start = report_date - timedelta(days=10)
            current_dates = [d.date() for d in pd.date_range(curr_start, periods=3)]
            prev_dates = [d.date() for d in pd.date_range(prev_start, periods=3)]
            label_curr, label_prev = "本期(周五~周日)", "上期(上周周五~周日)"
        else:
            # 周二~周五：本期 = 昨日, 上期 = 上周同期
            current_dates = [(report_date - timedelta(days=1)).date()]
            prev_dates = [(report_date - timedelta(days=8)).date()]
            label_curr, label_prev = "昨日", "上周同期"
        
        # 拉取数据
        metrics = [
            "曝光人数","访问人数","购买人数","消费金额",
            "新好评数","新中差评数","打卡人数","扫码人数","新增收藏人数","点评星级"
        ]
        cols = ", ".join(f"`{m}`" for m in metrics)
        cd_sql = ", ".join(f"'{d}'" for d in current_dates)
        pd_sql = ", ".join(f"'{d}'" for d in prev_dates)
        
        df_curr = pd.read_sql(
            f"SELECT `日期`, {cols} FROM `operation_data` "
            f"WHERE `美团门店ID` IN ({mids_sql}) AND `日期` IN ({cd_sql})",
            engine
        ).set_index("日期")
        
        df_prev = pd.read_sql(
            f"SELECT `日期`, {cols} FROM `operation_data` "
            f"WHERE `美团门店ID` IN ({mids_sql}) AND `日期` IN ({pd_sql})",
            engine
        ).set_index("日期")
        
        # 汇总并构造对比表
        if df_curr.empty or df_prev.empty:
            logger.warning(f"⚠️ 品牌 {brand} 本期或上期无数据，跳过对比表")
            return
            
        curr_sum = df_curr.sum()
        prev_sum = df_prev.sum()
        df_cmp = pd.DataFrame({
            "指标": metrics,
            label_curr: [curr_sum[m] for m in metrics],
            label_prev: [prev_sum[m] for m in metrics],
        })
        df_cmp["变化率"] = (
            (df_cmp[label_curr] - df_cmp[label_prev]) / df_cmp[label_prev] * 100
        ).round(1).astype(str) + "%"
        
        # 展示并保存
        fig, ax = plt.subplots(figsize=(len(df_cmp)*0.8, 2.5))
        ax.axis("off")
        tbl = ax.table(
            cellText=df_cmp.values,
            colLabels=df_cmp.columns,
            loc="center"
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(12)
        tbl.scale(1, 1.5)
        ax.set_title(brand, fontsize=16, pad=10)
        plt.tight_layout()
        
        # 确保输出目录存在
        out_dir = Path(REPORT_OUTPUT_DIR)
        out_dir.mkdir(exist_ok=True)
        out_path = out_dir / f"{brand}_同期对比表.png"
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        logger.info(f"✅ 保存对比表：{out_path}")
        
    except ImportError:
        logger.warning("matplotlib未安装，跳过对比表生成")
    except Exception as e:
        logger.error(f"生成对比表失败: {e}")
    """

def generate_daily_report(report_date, output_dir):
    """
    生成日报并输出到指定目录
    :param report_date: 报告日期
    :param output_dir: 输出目录
    """
    logger = get_logger('generate_daily_report')
    
    # 确保输出目录存在
    out_dir = Path(output_dir)
    out_dir.mkdir(exist_ok=True)
    
    # 获取门店映射
    store_map = get_store_mapping()
    if store_map.empty:
        logger.error("无法获取门店映射数据，退出")
        return
    
    # 计算日期范围
    last_7_start = report_date - timedelta(days=7)
    last_14_start = report_date - timedelta(days=14)
    
    logger.info(f"开始生成 {report_date.date()} 的日报")
    
    try:
        # 拉取昨日运营数据
        logger.info("拉取昨日运营数据")
        op_today = pd.read_sql(
            f"""SELECT 
                   `美团门店ID`,`曝光人数`,`访问人数`,`购买人数`,
                   `消费金额`,`新好评数`,`新中差评数`,
                   `打卡人数`,`扫码人数`,`点评星级`,`新增收藏人数`,`rankings_detail`
               FROM `operation_data`
               WHERE `日期` = '{report_date.date()}'""",
            engine
        ).merge(
            store_map[['美团门店ID','brand_name','operator']],
            on='美团门店ID', how='left'
        ).rename(columns={'brand_name':'推广门店'})

        # 拉取近7天运营数据
        logger.info("拉取近7天运营数据")
        op_last7 = pd.read_sql(
            f"""SELECT
                   `美团门店ID`,`日期`,`曝光人数`,`访问人数`,`购买人数`,
                   `消费金额`,`新好评数`,`新中差评数`,
                   `打卡人数`,`扫码人数`,`点评星级`,`新增收藏人数`
               FROM `operation_data`
               WHERE `日期` BETWEEN '{last_7_start.date()}' AND '{report_date.date()}'""",
            engine
        ).merge(
            store_map[['美团门店ID','brand_name','operator']],
            on='美团门店ID', how='left'
        ).rename(columns={'brand_name':'推广门店'})

        # 拉取昨日CPC数据
        logger.info("拉取昨日CPC数据")
        cpc_today = pd.read_sql(
            f"""SELECT `store_id`,`cost`,`impressions`,`clicks`,`orders`
               FROM `cpc_hourly_data`
               WHERE `date` = '{report_date.date()}'""",
            engine
        ).merge(
            store_map[['store_id','brand_name','operator']],
            on='store_id', how='left'
        ).rename(columns={'brand_name':'推广门店'})

        # 拉取最近14天历史数据
        logger.info("拉取最近14天运营数据")
        op_hist = pd.read_sql(
            f"""SELECT 
                   `美团门店ID`,`日期`,`曝光人数`,`访问人数`,`购买人数`,
                   `消费金额`,`新好评数`,`新中差评数`,
                   `打卡人数`,`扫码人数`,`点评星级`,`新增收藏人数`
               FROM `operation_data`
               WHERE `日期` BETWEEN '{last_14_start.date()}' AND '{report_date.date()}'""",
            engine
        ).merge(
            store_map[['美团门店ID','brand_name','operator']],
            on='美团门店ID', how='left'
        ).rename(columns={'brand_name':'推广门店'})

        # 分组准备
        op_group = op_today.groupby("推广门店")
        op7_group = op_last7.groupby("推广门店")
        cpc_group = cpc_today.groupby("推广门店")

        # ==================== 图表生成调用（已注释） ====================
        # TODO: 用户要求暂时忽略图片功能，后续会大改
        # 如需重新启用，请取消以下注释
        """
        # 先为每家门店生成对比图表
        logger.info("开始生成图表...")
        for brand in op_group.groups.keys():
            plot_vertical_table(brand, report_date, engine)
            plot_comparison_table(brand, report_date)
        """

        # 按运营师收集报告内容
        from collections import defaultdict
        operator_sections = defaultdict(list)

        # 指标字段
        op_fields = ["曝光人数","访问人数","购买人数","消费金额",
                    "新好评数","新中差评数","打卡人数","扫码人数","点评星级","新增收藏人数"]
        cpc_fields = ["cost","impressions","clicks","orders"]

        # 为每个品牌生成报告
        for brand, df_op in op_group:
            # 准备当日7天历史 & CPC & 全量历史数据
            df_op7 = op7_group.get_group(brand) if brand in op7_group.groups else pd.DataFrame()
            df_cpc = cpc_group.get_group(brand) if brand in cpc_group.groups else pd.DataFrame()
            df_hist_b = op_hist[op_hist["推广门店"] == brand]

            # 1) 汇总当日运营 & CPC 数据
            op_sum = summarize(df_op, op_fields)
            if df_cpc.empty:
                cpc_sum, ratios = {}, {}
            else:
                cpc_sum = summarize(df_cpc, cpc_fields)
                ratios = compute_cpc_contribution_ratios(op_sum, cpc_sum)

            # 2) 计算日环比（昨日 vs. 上周同期），周一特殊处理
            weekday = report_date.weekday()
            if weekday == 0:
                # 周一：把 curr 定义为上周五~周日，prev 定义为上上周五~周上周日
                curr = df_hist_b[
                    (df_hist_b["日期"] >= (report_date - timedelta(days=3)).date()) &
                    (df_hist_b["日期"] <= (report_date - timedelta(days=1)).date())
                ]
                prev = df_hist_b[
                    (df_hist_b["日期"] >= (report_date - timedelta(days=10)).date()) &
                    (df_hist_b["日期"] <= (report_date - timedelta(days=8)).date())
                ]
                curr_sum = summarize(curr, op_fields)
            else:
                # 非周一：本期就是昨天，prev 是上周同一天
                curr = df_hist_b[df_hist_b["日期"] == (report_date - timedelta(days=1)).date()]
                prev = df_hist_b[df_hist_b["日期"] == (report_date - timedelta(days=8)).date()]
                curr_sum = op_sum

            prev_sum = summarize(prev, op_fields)

            # 组装环比数字
            link_ratio = {}
            for k in op_fields:
                if prev_sum.get(k, 0):
                    val_curr = curr_sum.get(k, 0)
                    val_prev = prev_sum.get(k, 0)
                    pct = round((val_curr - val_prev) / (val_prev or 1) * 100, 1)
                    link_ratio[k] = f"{pct}%"
                else:
                    link_ratio[k] = "N/A"

            # 3) 解析榜单动态
            non_null = df_op["rankings_detail"].dropna().tolist() if not df_op.empty else []
            raw_rank = non_null[0] if non_null else None
            if isinstance(raw_rank, (bytes, bytearray)):
                try: 
                    raw_rank = raw_rank.decode("utf-8")
                except: 
                    raw_rank = None
            if isinstance(raw_rank, str):
                try: 
                    parsed = json.loads(raw_rank)
                except: 
                    parsed = {}
            elif isinstance(raw_rank, dict):
                parsed = raw_rank
            else:
                parsed = {}
            if isinstance(parsed, str):
                try: 
                    tmp = json.loads(parsed)
                except: 
                    tmp = {}
                parsed = tmp if isinstance(tmp, dict) else {}
            rank_dict = parsed if isinstance(parsed, dict) else {}

            # 只保留这几个平台
            allowed = ["dianping_hot", "dianping_checkin", "dianping_rating"]
            th_city = 10
            th_sub = 5
            level_map = {"city": "全市榜", "subdistrict": "区县榜", "business": "商圈榜"}

            rank_lines = []
            for pf in allowed:
                scope = rank_dict.get(pf, {})
                if not isinstance(scope, dict):
                    continue
                desc = RANK_PLATFORMS[pf]
                city_rank = scope.get("city")
                sub_rank = scope.get("subdistrict")
                bus_rank = scope.get("business")
                
                if isinstance(city_rank, int) and city_rank <= th_city:
                    rank_lines.append(f"{desc}{level_map['city']}第{city_rank}名")
                if isinstance(sub_rank, int) and sub_rank <= th_sub:
                    rank_lines.append(f"{desc}{level_map['subdistrict']}第{sub_rank}名")
                if not (
                    (isinstance(city_rank, int) and city_rank <= th_city) or
                    (isinstance(sub_rank, int) and sub_rank <= th_sub)
                ) and isinstance(bus_rank, int):
                    rank_lines.append(f"{desc}{level_map['business']}第{bus_rank}名")

            rank_note = "\n".join(rank_lines) if rank_lines else "无榜单变化"

            # 4) 取昨日四项核心指标 & 环比
            today = df_op.iloc[0]
            card_val = int(today['打卡人数'])
            good_val = int(today['新好评数'])
            bad_val = int(today['新中差评数'])
            rev_val = round(today['消费金额'], 0)

            card_pct = link_ratio.get('打卡人数', 'N/A')
            good_pct = link_ratio.get('新好评数', 'N/A')
            bad_pct = link_ratio.get('新中差评数', 'N/A')
            rev_pct = link_ratio.get('消费金额', 'N/A')

            # 格式化输出
            card_str = fmt(card_val, card_pct)
            good_str = fmt(good_val, good_pct)
            bad_str = "暂无差评" if bad_val == 0 else fmt(bad_val, bad_pct)
            rev_str = fmt(rev_val, rev_pct)

            # 新增收藏
            col_val = int(op_sum.get("新增收藏人数", 0))
            col_pct = link_ratio.get("新增收藏人数", "N/A")
            col_str = fmt(col_val, col_pct)

            # 异动检测
            extra_notes = []
            exp_pct = link_ratio.get("曝光人数", "N/A")
            if is_big_change(exp_pct):
                extra_notes.append(f"⚠️ 曝光{exp_pct}")
            vis_pct = link_ratio.get("访问人数", "N/A")
            if is_big_change(vis_pct):
                extra_notes.append(f"⚠️ 访问{vis_pct}")
            extra_note_str = "；" + "；".join(extra_notes) if extra_notes else ""

            # 构建报告文本
            suggestion = ("发现差评，运营师已介入跟进。" if bad_val > 0 else "数据正常，继续引导好评。")
            
            # 推广通花费分析
            cpc_part = ""
            cost_today = cpc_sum.get("cost", 0) if not df_cpc.empty else 0
            cost_prev = 0  # 这里可以扩展为获取上周同期CPC数据
            if not (cost_today == 0 and cost_prev == 0):
                cpc_part = f"；推广通花费 {cost_today:.2f}"
                # 判断推广通花费异常
                if (cost_prev > 0 and cost_today == 0) or (
                    cost_prev > 0 and abs(cost_today - cost_prev)/cost_prev >= 0.5
                    and abs(cost_today - cost_prev) >= 100
                ):
                    cpc_part += " ⚠️ 推广通花费异常"
            
            # 数据价值分析
            insights = analyze_brand_performance(brand, df_op, op_sum, link_ratio)
            insights_text = "\n".join(insights) if insights else "数据表现正常"
            
            text = (
                f"{brand}\n"
                f"- 核心指标：消费 {rev_str}{cpc_part}；打卡 {card_str}；收藏 {col_str}；\n"
                f"  好评 {good_str}；{'' if bad_val == 0 else '差评 ' + bad_str}\n"
                f"- 排行榜：\n{rank_note}\n"
                f"- 数据洞察：\n{insights_text}\n"
                f"- 建议：{suggestion}\n"
            )

            # 按运营师收集
            match = store_map[store_map['brand_name'] == brand]
            if match.empty:
                logger.warning(f"品牌 {brand} 未在 store_map 中匹配到，跳过")
                continue
            op_val = match['operator'].values[0]
            if pd.isna(op_val) or not op_val:
                logger.warning(f"品牌 {brand} 找到了但运营师字段为空，跳过")
                continue

            logger.info(f"品牌 {brand} 匹配运营师：{op_val}")
            operator_sections[op_val].append(text)

        # 写各运营师的 TXT 文件
        for op, texts in operator_sections.items():
            fn = out_dir / f"{op}_{report_date.date()}.txt"
            fn.write_text("\n\n".join(texts), encoding="utf-8-sig")
            logger.info(f"已生成运营师日报：{fn}")

        # 生成Excel文件
        logger.info("开始生成Excel月度数据文件")
        
        # 拉取当月全量数据
        month_start = report_date.replace(day=1).date()
        df_month = pd.read_sql(
            f"""
            SELECT
              `美团门店ID`,`日期`,`消费金额`,`曝光人数`,`访问人数`,`购买人数`,
              `扫码人数`,`新增收藏人数`,`打卡人数`,`新好评数`,`新中差评数`,`点评星级`
            FROM `operation_data`
            WHERE `日期` BETWEEN '{month_start}' AND '{report_date.date()}'
            """, engine
        ).merge(
            store_map[['美团门店ID','brand_name','operator','store_id']],
            on='美团门店ID', how='left'
        )

        # 拉取当月每日CPC成本
        cpc_month = pd.read_sql(
            f"""
            SELECT 
              store_id, `date` AS 日期, SUM(cost) AS 推广通花费
            FROM cpc_hourly_data
            WHERE `date` BETWEEN '{month_start}' AND '{report_date.date()}'
            GROUP BY store_id, `date`
            """, engine
        ).merge(
            store_map[['store_id','brand_name','operator']],
            on='store_id', how='left'
        )
        
        # 同一天同品牌可能存在多门店，按品牌+日期汇总
        cpc_month = cpc_month.groupby(
            ['brand_name','operator','日期'], as_index=False
        )['推广通花费'].sum()

        # 把推广通花费合并回df_month，缺失时设为0
        df_month = df_month.merge(
            cpc_month[['brand_name','日期','推广通花费']],
            on=['brand_name','日期'], how='left'
        ).fillna({'推广通花费': 0})
        df_month['推广通花费'] = df_month['推广通花费'].round(2)

        # 衍生"星期"、"访问转化"、"购买转化"
        weekday_map = {0:'星期一',1:'星期二',2:'星期三',3:'星期四',4:'星期五',5:'星期六',6:'星期日'}
        df_month['星期'] = pd.to_datetime(df_month['日期']).dt.weekday.map(weekday_map)
        df_month['访问转化'] = (df_month['访问人数']/df_month['曝光人数']).round(3)
        df_month['购买转化'] = (df_month['购买人数']/df_month['访问人数']).round(3)

        # 定义列顺序
        cols = [
            '日期','星期','消费金额','推广通花费','曝光人数','访问人数','购买人数',
            '访问转化','购买转化','新增收藏人数','打卡人数',
            '新好评数','新中差评数','扫码人数','点评星级'
        ]

        # 按运营师输出月度Excel（每店一个sheet）
        for op, grp_op in df_month.groupby('operator'):
            file = out_dir / f"{op}_{report_date.date()}_月度数据.xlsx"
            
            # 写每个门店分别一个sheet
            with pd.ExcelWriter(file, engine='openpyxl', mode='w') as writer:
                for brand, grp in grp_op.groupby('brand_name'):
                    sheet = grp.sort_values('日期')[cols]
                    sheet.to_excel(
                        writer,
                        sheet_name=brand[:31],  # sheet名称截31字
                        index=False,
                        startrow=2
                    )
            logger.info(f"已生成运营师月度Excel：{file}")

            # 样式设置
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.load_workbook(file)
                header_fill = PatternFill("solid", fgColor="DDDDDD")
                
                for ws in wb.worksheets:
                    max_col = ws.max_column

                    # 合并第1行写门店名称
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                    title = ws.cell(row=1, column=1)
                    title.value = ws.title
                    title.font = Font(size=14, bold=True)
                    title.alignment = Alignment(horizontal="center", vertical="center")

                    # 第3行样式：加粗居中＋灰底
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=3, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # 冻结前三行
                    ws.freeze_panes = "A4"

                    # 列宽自适应 & 强制最小宽度
                    min_widths = {
                        'A': 13,  # 日期
                        'B': 8,   # 星期
                        'C': 10,  # 消费金额
                        'D': 10, 'E': 10, 'F': 10, 'G': 10,
                        'H': 10, 'I': 13, 'J': 10, 'K': 10, 'L': 11,
                        'M': 8,   # 扫码
                        'N': 8    # 星级
                    }
                    
                    for col_cells in ws.columns:
                        col_letter = get_column_letter(col_cells[0].column)
                        if col_letter == 'A':
                            ws.column_dimensions[col_letter].width = min_widths['A']
                            continue

                        max_len = max(
                            len(str(c.value)) if c.value is not None else 0
                            for c in col_cells
                        )
                        optimal = max_len + 4
                        ws.column_dimensions[col_letter].width = max(
                            optimal, min_widths.get(col_letter, optimal)
                        )

                    # 单元格格式化
                    fmt_amt2 = '#,##0.00'  # 两位小数
                    fmt_int = '#,##0'      # 整数
                    pct_fmt = '0.0%'       # 百分比
                    star_fmt = '0.0'       # 星级一位小数
                    
                    for row in range(4, ws.max_row + 1):
                        # C列：消费金额保留两位小数
                        ws.cell(row, 3).number_format = fmt_amt2
                        # D列：推广通花费保留两位小数
                        ws.cell(row, 4).number_format = fmt_amt2
                        # E~G列：曝光/访问/购买用整数
                        ws.cell(row, 5).number_format = fmt_int
                        ws.cell(row, 6).number_format = fmt_int
                        ws.cell(row, 7).number_format = fmt_int
                        # H列：访问转化百分比
                        ws.cell(row, 8).number_format = pct_fmt
                        # I列：购买转化百分比
                        ws.cell(row, 9).number_format = pct_fmt
                        # 最后一列（点评星级）用一位小数
                        ws.cell(row, ws.max_column).number_format = star_fmt

                wb.save(file)
                logger.info(f"Excel已格式化：{file}")
                
            except ImportError:
                logger.warning("openpyxl未安装，跳过Excel样式设置")
            except Exception as e:
                logger.error(f"Excel样式设置失败: {e}")

        logger.info("日报生成完成")
        
    except Exception as e:
        logger.error(f"生成日报时发生错误: {e}")
        raise

def main():
    """主函数"""
    logger = get_logger('generate_daily_report')
    
    # 默认使用昨天的日期
    default_dt = datetime.now() - timedelta(days=1)
    default_str = default_dt.strftime("%Y-%m-%d")
    
    # 交互式日期选择
    import sys
    if len(sys.argv) > 1:
        date_str = sys.argv[1]
    else:
        # 交互式输入日期
        inp = input(f"使用日期 [{default_str}]？回车确认 或 输入其他日期 (YYYY-MM-DD)：").strip()
        date_str = inp if inp else default_str
    
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        logger.error(f"日期格式错误：{date_str}，请使用 YYYY-MM-DD 格式")
        return
    
    logger.info(f"🗓️ 最终使用日期：{report_date.date()}，生成日报")
    logger.info(f"日报将输出到: {REPORT_OUTPUT_DIR}")
    
    generate_daily_report(report_date, REPORT_OUTPUT_DIR)
    logger.info("日报流程已完成")

if __name__ == '__main__':
    main()
