import os
import pandas as pd
from sqlalchemy import create_engine, text
import sqlalchemy.types as sqltypes
from excel_header_finder import clean_and_load_excel
from data_cleaning import clean_numeric_columns
from database_importer import import_to_mysql
import send2trash
from config import DB_CONNECTION_STRING

# 商品日明细自动导入与月报生成管道示例

# === 配置区域 ===
PRODUCT_DAILY_FOLDER = r"D:\dianping_downloads\product_daily_data"
TEMPLATE_PATH = r"monthly_report_template.xlsx"  # 请准备含占位符的Excel模板

# === 商品数据导入 ===
def delete_old_product(start_date, end_date):
    """
    删除数据库中指定日期区间的历史商品日明细，确保导入时不重复。"""
    engine = create_engine(DB_CONNECTION_STRING)
    with engine.begin() as conn:
        res = conn.execute(
            text("DELETE FROM product_daily WHERE date BETWEEN :s AND :e"),
            {"s": start_date, "e": end_date}
        )
        print(f"✅ 删除商品日明细 {start_date}~{end_date} 共 {res.rowcount} 条。")


def process_product_folder():
    """
    扫描 PRODUCT_DAILY_FOLDER 下所有.xlsx文件，清洗并导入到 product_daily 表。
    导入后将源文件移入回收站。
    """
    engine = create_engine(DB_CONNECTION_STRING)
    dfs, filepaths = [], []

    for fname in os.listdir(PRODUCT_DAILY_FOLDER):
        if not fname.lower().endswith('.xlsx'):
            continue
        fp = os.path.join(PRODUCT_DAILY_FOLDER, fname)
        df = clean_and_load_excel(fp)
        df = clean_numeric_columns(df, key_col='商品ID')
        df['日期'] = pd.to_datetime(df['日期'], errors='coerce').dt.date
        df.drop_duplicates(subset=['日期', '商品ID'], keep='last', inplace=True)
        dfs.append(df)
        filepaths.append(fp)

    if not dfs:
        print("⚠️ 无商品数据文件可导入")
        return

    df_all = pd.concat(dfs, ignore_index=True)
    min_date, max_date = df_all['日期'].min(), df_all['日期'].max()

    delete_old_product(min_date, max_date)

    # 定义数据库字段类型映射，可根据实际列补充
    dtype_map = {
        '日期': sqltypes.Date,
        '商品ID': sqltypes.String(64),
        '商品名称': sqltypes.String(255),
        '商品访问人数': sqltypes.Integer,
        '商品购买人数': sqltypes.Integer,
        '商品成交金额(优惠后)': sqltypes.Float
    }

    import_to_mysql(df_all, 'product_daily', DB_CONNECTION_STRING, dtype=dtype_map)
    print(f"✅ 成功导入商品日明细，共 {len(df_all)} 行。")

    for fp in filepaths:
        send2trash.send2trash(fp)
        print(f"🗑️ 已移入回收站：{fp}")


# === 月报生成 ===
def generate_monthly_report(start_date: str, end_date: str, template_path: str, output_path: str):
    """
    从数据库读取运营、CPC、商品数据，计算主要指标并填充到Excel模板，输出月报文件。
    """
    engine = create_engine(DB_CONNECTION_STRING)

    # 1. 读取数据
    ops = pd.read_sql(text(
        "SELECT * FROM operation_data WHERE `日期` BETWEEN :s AND :e"
    ), engine, params={'s': start_date, 'e': end_date})

    cpcs = pd.read_sql(text(
        "SELECT date, store_id, cost, impressions, clicks FROM cpc_hourly_data WHERE date BETWEEN :s AND :e"
    ), engine, params={'s': start_date, 'e': end_date})

    prods = pd.read_sql(text(
        "SELECT * FROM product_daily WHERE date BETWEEN :s AND :e"
    ), engine, params={'s': start_date, 'e': end_date})

    # 2. 计算指标示例
    total_gmv = ops['成交金额(优惠后)'].sum()
    avg_gmv = total_gmv / ops['日期'].nunique()
    peak_day = ops.groupby('日期')['成交金额(优惠后)'].sum().idxmax()

    total_cost = cpcs['cost'].sum()
    avg_cpc = total_cost / cpcs['clicks'].sum() if cpcs['clicks'].sum() > 0 else 0

    top_products = prods.groupby('商品名称').agg({
        '商品购买人数': 'sum',
        '商品访问人数': 'sum',
        '商品成交金额(优惠后)': 'sum'
    }).sort_values('商品购买人数', ascending=False).head(10)

    # 3. 填充 Excel
    from openpyxl import load_workbook
    wb = load_workbook(template_path)

    # 示例：填充【经营概览】工作表
    ws = wb['经营概览']
    ws['B2'] = total_gmv
    ws['B3'] = avg_gmv
    ws['B4'] = peak_day.strftime('%Y-%m-%d')

    # 示例：填充【推广分析】
    ws2 = wb['推广分析']
    ws2['B2'] = total_cost
    ws2['B3'] = avg_cpc

    # 示例：填充【商品分析】Top10
    ws3 = wb['商品分析']
    for idx, (name, row) in enumerate(top_products.iterrows(), start=2):
        ws3.cell(row=idx, column=1, value=name)
        ws3.cell(row=idx, column=2, value=row['商品购买人数'])
        ws3.cell(row=idx, column=3, value=row['商品访问人数'])
        ws3.cell(row=idx, column=4, value=row['商品成交金额(优惠后)'])

    # 保存输出
    wb.save(output_path)
    print(f"✅ 月报已生成：{output_path}")


if __name__ == '__main__':
    # 运行示例：先导入商品数据，再生成7月报告
    process_product_folder()
    generate_monthly_report(
        start_date='2025-07-01',
        end_date='2025-07-31',
        template_path=TEMPLATE_PATH,
        output_path='monthly_report_2025_07.xlsx'
    )
