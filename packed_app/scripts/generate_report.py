"""
自动生成周报/月报（首版骨架）：

流程：
1) 读取配置与门店映射（store_mapping） → 品牌、门店、分店名
2) 用 Playwright + 已登录的持久化浏览器打开页面：
   - 交易概览页：若“已选全部”，根据门店ID筛选后确认
   - 截取核心指标区域截图
   - 流量概览页：抓取“流失去向-同行 Top3”
   - 推广分析页：抓取花费、排名等
3) 读取数据库汇总目标周期指标，计算环比与增减量（万为单位格式化）
4) 写入 Excel：标题、截图、要点描述与数字

说明：首版先打通全链路，选择器/格式化细节按照实际页面需要可继续调优。
"""

from __future__ import annotations

import os
import io
import math
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import List, Tuple, Dict, Optional
import re

import yaml
import pandas as pd
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

from config.config import DATA_DOWNLOAD_DIR
from app.db import get_engine, reflect_existing_columns
from scripts.download_common import open_chromium_context, close_all
from scripts.profiles import PROFILE_BRAND_MAP


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def load_cfg(path: Path) -> dict:
    with open(path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def get_cycle_dates(kind: str = 'week') -> Tuple[date, date]:
    today = date.today()
    if kind == 'week':
        end = today - timedelta(days=today.weekday() + 1)  # 默认上周日
        start = end - timedelta(days=6)
    else:
        # 默认上月
        first_this = today.replace(day=1)
        end = first_this - timedelta(days=1)
        start = end.replace(day=1)
    return start, end


def humanize_number(n: float) -> str:
    try:
        n = float(n)
    except Exception:
        return str(n)
    if abs(n) >= 10000:
        return f"{n/10000:.2f}w"
    if abs(n) >= 1000:
        return f"{n/1000:.1f}k"
    if n.is_integer():
        return f"{int(n)}"
    return f"{n:.2f}"


def fetch_store_mapping() -> pd.DataFrame:
    eng = get_engine()
    return pd.read_sql("SELECT * FROM store_mapping", eng)


def _try_click(root, selector: str, timeout: int = 2000) -> bool:
    try:
        root.locator(selector).first.click(timeout=timeout)
        return True
    except Exception:
        return False


def _get_main_frame(page):
    try:
        return page.frame_locator("iframe").first
    except Exception:
        return None


def ensure_store_selected(page, cfg: dict, store_id: str, store_name: Optional[str] = None) -> bool:
    sel = cfg['selectors']
    frame = _get_main_frame(page)
    # 若页面只有一个门店或已选定，则跳过
    try:
        val = (frame.get_by_placeholder("全部门店").input_value(timeout=1200) if frame else None)
        need_pick = ('已选全部' in (val or ''))
    except Exception:
        try:
            val = page.locator("xpath=" + sel['store_input_value']).input_value(timeout=1200)
            need_pick = ('已选全部' in val)
        except Exception:
            need_pick = False
    if not need_pick:
        return True
    # 打开弹框（优先点 iframe 内的占位符输入框）
    if frame:
        try:
            frame.get_by_placeholder("全部门店").click()
        except Exception:
            _try_click(frame, "xpath=" + sel['store_input_value'])
    else:
        _try_click(page, "xpath=" + sel['store_input_value'])
    # 先取消已有选择（默认全选）——优先用用户提供的绝对 XPath
    try:
        page.locator("xpath=" + sel['clear_selected_xpath']).click(timeout=1500)
    except Exception:
        try:
            page.get_by_text("清空已选", exact=True).click(timeout=1500)
        except Exception:
            pass
    # 清空后点击“确定”（第2个按钮），应用清空但不关闭弹窗
    try:
        page.get_by_role("button", name="确定").nth(1).click(timeout=1500)
    except Exception:
        pass
    # 勾选（优先按“门店ID”精确匹配；若失败再按完整门店名；最后关键字）
    picked = False
    if frame:
        # 1) 用 ID 精确匹配
        if store_id and not picked:
            try:
                row_xpath = sel['store_table_row_by_id_tpl'].format(store_id=store_id)
                row = page.locator("xpath=" + row_xpath).first
                row.wait_for(timeout=3000)
                row.get_by_role("cell").first.click(timeout=2000)
                picked = True
            except Exception:
                picked = False
        # 2) 完整门店名
        if store_name and not picked:
            try:
                full_row = page.get_by_role("row", name=re.compile(re.escape(str(store_name))))
                full_row.nth(0).get_by_role("cell").first.click(timeout=2000)
                picked = True
            except Exception:
                pass
        # 3) 关键字（括号内差异）
        if not picked:
            kw = None
            if store_name:
                m = re.search(r'[（(]([^（）()]+)[）)]', str(store_name))
                kw = (m.group(1) if m else None)
            if kw:
                try:
                    # 右侧搜索过滤
                    inp = frame.get_by_placeholder("门店ID/门店名称/城市")
                    inp.fill("")
                    frame.wait_for_timeout(200)
                    inp.fill(kw)
                    frame.wait_for_timeout(400)
                except Exception:
                    pass
                try:
                    name_row = page.get_by_role("row", name=re.compile(re.escape(kw)))
                    name_row.nth(0).get_by_role("cell").first.click(timeout=2000)
                    picked = True
                except Exception:
                    picked = False
    # 若存在“全选”开关，先取消全选（避免多店）
    if frame:
        try:
            frame.get_by_label("全选").uncheck(timeout=800)
        except Exception:
            pass
    # 最终确认（优先用绝对 XPath），关闭弹窗
    try:
        page.locator("xpath=" + sel['store_modal_confirm_xpath']).click(timeout=1500)
    except Exception:
        try:
            page.get_by_role("button", name="确定").last.click(timeout=1500)
        except Exception:
            if frame:
                try:
                    frame.get_by_role("button", name="确定").click()
                except Exception:
                    _try_click(frame, "xpath=" + sel['store_confirm_btn'])
            else:
                _try_click(page, "xpath=" + sel['store_confirm_btn'])
    # 等待输入框 value 变化（避免 networkidle 卡住）
    ok = False
    for _ in range(20):  # ~6s
        try:
            val2 = (frame.get_by_placeholder("全部门店").input_value(timeout=300) if frame else page.locator("xpath=" + sel['store_input_value']).input_value(timeout=300))
            if val2 and ('已选全部' not in val2):
                ok = True
                break
        except Exception:
            pass
        page.wait_for_timeout(300)
    return ok or picked


def screenshot_core_box(page, cfg: dict, save_path: Path, allow_fallback: bool = True) -> List[Path]:
    box_xpath = cfg['selectors']['trade_core_box']
    scale = cfg['screenshot'].get('scale', 'css')
    full_fallback = bool(cfg['screenshot'].get('full_page_fallback', True))
    frame = _get_main_frame(page)
    # 为了“从页面上方到表格下方”的大范围截图：
    # 1) 找到顶部锚点，滚动到顶部
    try:
        top_anchor = (frame.locator("xpath=" + cfg['selectors']['page_top_anchor']).first) if frame else page.locator("xpath=" + cfg['selectors']['page_top_anchor']).first
        top_anchor.scroll_into_view_if_needed(timeout=1000)
    except Exception:
        pass
    box = (frame.locator("xpath=" + box_xpath).first) if frame else page.locator("xpath=" + box_xpath).first
    try:
        box.wait_for(state='visible', timeout=5000)
        # 2) 分两屏截图：
        # 顶部一屏
        page.evaluate("window.scrollTo(0, 0)")
        page.wait_for_timeout(400)
        img_paths: List[Path] = []
        p1 = save_path.with_name(save_path.stem + "_p1" + save_path.suffix)
        page.screenshot(path=str(p1), full_page=False, scale=scale)
        img_paths.append(p1)
        # 内部 iframe 下滚一屏，再截第二屏
        try:
            iframe_el = page.query_selector("iframe")
            if iframe_el is not None:
                inner_frame = iframe_el.content_frame()
                if inner_frame is not None:
                    inner_frame.evaluate("window.scrollBy(0, window.innerHeight)")
        except Exception:
            # 退回使用 Page 滚动
            page.evaluate("window.scrollBy(0, window.innerHeight)")
        page.wait_for_timeout(500)
        p2 = save_path.with_name(save_path.stem + "_p2" + save_path.suffix)
        page.screenshot(path=str(p2), full_page=False, scale=scale)
        img_paths.append(p2)
        return img_paths
    except Exception:
        if allow_fallback and full_fallback:
            # Fallback：仍然返回单张全页
            page.screenshot(path=str(save_path), full_page=True, scale=scale)
            return [save_path]
        else:
            raise


def select_week_and_trend(page, cfg: dict, start: date, end: date) -> bool:
    sel = cfg['selectors']
    frame = _get_main_frame(page)
    if frame is None:
        return False
    # 点击“单周”
    success = True
    try:
        try:
            frame.get_by_role("button", name="单周").click()
        except Exception:
            if not _try_click(frame, "xpath=" + sel['week_button']):
                _try_click(frame, "button:has-text(\u5355\u5468)")
    except Exception:
        success = False
    # 打开周选择器
    try:
        week_days = " ".join(str(d) for d in range(start.day, start.day + 7))
        try:
            frame.get_by_role("button", name=week_days).click()
        except Exception:
            _try_click(frame, "xpath=" + sel['datepicker_trigger'])
            row = frame.locator("xpath=" + sel['week_row_tpl'].format(week_no=end.isocalendar()[1])).first
            row.click(timeout=3000)
    except Exception:
        success = False
    # 点击“店均及同行趋势”
    try:
        try:
            frame.get_by_role("button", name="店均及同行趋势").click()
        except Exception:
            if not _try_click(frame, "xpath=" + sel['trend_button']):
                _try_click(frame, "button:has-text(\u5e97\u5747\u53ca\u540c\u884c\u8d8b\u52bf)")
    except Exception:
        success = False
    # 等待图表渲染：先等图例出现，再额外等待一段时间
    try:
        for legend in ["我的店均-当前", "我的店均-往年同期", "商圈同行店均", "商圈标杆店均"]:
            try:
                frame.get_by_text(legend, exact=True).wait_for(timeout=2500)
            except Exception:
                pass
        frame.wait_for_timeout(2500)
    except Exception:
        pass
    return success


def query_operation_kpis(store_mt_id: str, start: date, end: date) -> Dict[str, float]:
    """从 operation_data 读取聚合指标，自动适配不同列名变体。"""
    eng = get_engine()
    existing = set(reflect_existing_columns("operation_data"))

    # 为每个指标定义候选列，按优先级匹配
    col_candidates: Dict[str, List[str]] = {
        'expo': ['曝光人数', '曝光次数'],
        'visit': ['访问人数', '访问次数'],
        'buy': ['购买人数'],
        'fav': ['收藏（次）', '新增收藏人数', '收藏人数', '累计收藏人数'],
        'checkin': ['打卡人数', '打卡'],
    }

    chosen: Dict[str, Optional[str]] = {k: next((c for c in v if c in existing), None) for k, v in col_candidates.items()}

    select_cols = ['`日期`'] + [f"`{c}`" for c in chosen.values() if c]
    sql = f"SELECT {', '.join(select_cols)} FROM operation_data WHERE `美团门店ID`=%s AND DATE(`日期`) BETWEEN %s AND %s"
    df = pd.read_sql(sql, eng, params=(store_mt_id, start, end))
    if df.empty:
        return {}

    def sum_col(name: Optional[str]) -> float:
        return float(df[name].astype(float).sum()) if name and name in df.columns else 0.0

    return {
        '曝光': sum_col(chosen['expo']),
        '访问': sum_col(chosen['visit']),
        '购买': sum_col(chosen['buy']),
        '收藏': sum_col(chosen['fav']),
        '打卡': sum_col(chosen['checkin']),
    }


def query_operation_kpis_prev(store_mt_id: str, start: date, end: date) -> Dict[str, float]:
    span = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=span - 1)
    return query_operation_kpis(store_mt_id, prev_start, prev_end)


def build_summary_text(cur: Dict[str, float], prev: Dict[str, float]) -> str:
    parts = []
    for key in ['曝光', '访问', '购买']:
        c = cur.get(key, 0.0)
        p = prev.get(key, 0.0)
        diff = c - p
        pct = (diff / p * 100.0) if p else 0.0
        parts.append(f"{key}：{humanize_number(c)}（{humanize_number(diff)}），环比{pct:.1f}%")
    # 收藏/打卡
    sc = cur.get('收藏', 0.0); sp = prev.get('收藏', 0.0)
    dk = cur.get('打卡', 0.0); dp = prev.get('打卡', 0.0)
    sc_pct = ((sc-sp)/sp*100.0) if sp else 0.0
    dk_pct = ((dk-dp)/dp*100.0) if dp else 0.0
    parts.append(f"收藏打卡：收藏{sc_pct:+.2f}%，打卡{dk_pct:+.2f}%")
    return "\n".join(parts)


def write_excel(report_path: Path, title: str, summary: str, image_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = '周报'
    ws['A1'] = title
    ws['A3'] = summary
    # 插入图片（支持两屏 p1/p2 按顺序插入）
    p1 = image_path.with_name(image_path.stem + "_p1" + image_path.suffix)
    p2 = image_path.with_name(image_path.stem + "_p2" + image_path.suffix)
    anchor_row = 5
    def add_img_if_exists(p: Path):
        nonlocal anchor_row
        if p.exists():
            img = XLImage(str(p))
            ws.add_image(img, f'A{anchor_row}')
            anchor_row += 30  # 粗略下移，避免覆盖
    # 优先插入 p1/p2；若不存在则插入单张 image_path
    inserted = False
    for p in [p1, p2]:
        if p.exists():
            add_img_if_exists(p)
            inserted = True
    if not inserted and image_path.exists():
        add_img_if_exists(image_path)
    wb.save(str(report_path))


def run_store_in_page(page, settings: dict, brand: str, store_row: pd.Series, cycle: str = 'week') -> None:
    cfg = settings
    start, end = get_cycle_dates(cycle)
    report_dir = Path(cfg['excel']['output_dir'])
    brand_dir = report_dir / brand
    brand_dir.mkdir(parents=True, exist_ok=True)
    store_mt_id = str(store_row['美团门店ID'])
    store_name = str(store_row.get('推广门店') or store_row.get('门店名称') or brand)
    title = f"{store_name}{'周报' if cycle=='week' else '月报'}数据复盘 {start}~{end}"
    # 打开交易概览页 → 确保选择门店 → 切视图 → 截图
    page.goto(cfg['trade_overview_url'], wait_until='networkidle')
    ok_store = False
    try:
        ok_store = ensure_store_selected(page, cfg, store_mt_id, store_name)
    except Exception:
        ok_store = False
    # 打印当前“门店范围”输入框值，便于确认是否正确
    try:
        current_val = page.locator("xpath=" + cfg['selectors']['store_input_value']).input_value(timeout=1500)
        print(f"当前门店范围：{current_val}")
    except Exception:
        pass
    ok_view = False
    try:
        ok_view = select_week_and_trend(page, settings, start, end)
    except Exception:
        ok_view = False
    img_path = brand_dir / f"{store_name}_{cycle}_{start}_{end}.png"
    if ok_store and ok_view:
        try:
            screenshots = screenshot_core_box(page, cfg, img_path, allow_fallback=False)
            # 若有多张，记录第一张为主图
            if screenshots and screenshots[0] != img_path:
                pass
        except Exception as e:
            print(f"⚠️ 截图跳过（未达到指定视图或容器未找到）: {e}")
    else:
        print("⚠️ 未完成门店/单周/趋势设置，跳过截图")

    # 查询数据库 KPI 与环比
    cur = query_operation_kpis(store_mt_id, start, end)
    prev = query_operation_kpis_prev(store_mt_id, start, end)
    summary = build_summary_text(cur, prev)

    # 写入 Excel
    xls_path = brand_dir / f"{store_name}_{cycle}_{start}_{end}.xlsx"
    write_excel(xls_path, title, summary, img_path)
    print(f"✅ 报告完成: {xls_path}")


def main():
    settings = load_cfg(Path(__file__).with_name('report_settings.yaml'))
    # 复用下载流程的克隆目录与已有登录态
    # 从 settings.yaml 读取 user_data_dir 与 profiles。默认使用 'Profile 49' 对应品牌
    from scripts.download_data import load_settings
    base_dir = Path(__file__).parent
    cfg_dl = load_settings(base_dir / 'settings.yaml')
    user_data_dir = Path(cfg_dl['clone_dir']).expanduser().resolve()

    mapping = fetch_store_mapping()
    # 简化：只处理在 PROFILE_BRAND_MAP 中配置过的品牌对应的 profile
    brand_to_profile = {v['brand']: k for k, v in PROFILE_BRAND_MAP.items()}
    # 在一个持久上下文中循环多家门店，避免频繁启动浏览器
    for brand, prof in brand_to_profile.items():
        rows = mapping[mapping['推广门店'] == brand]
        if rows.empty:
            continue
        # 为该品牌使用对应的浏览器 Profile，确保打开的是该品牌已登录的账号
        p, ctx, page = open_chromium_context(user_data_dir, prof, Path(DATA_DOWNLOAD_DIR))
        try:
            for _, row in rows.iterrows():
                try:
                    run_store_in_page(page, settings, brand, row, cycle='week')
                except Exception as e:
                    print(f"❌ 生成 {brand}-{row['美团门店ID']} 失败: {e}")
        finally:
            close_all(p, ctx)


if __name__ == '__main__':
    main()


