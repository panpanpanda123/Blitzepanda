import argparse
import json
import pandas as pd
import time
from datetime import datetime, timedelta
from pathlib import Path
from requests.exceptions import SSLError

from config_and_brand import engine, API_KEY, MODEL, brand_profile
from AI_prompt import call_kimi_api, safe_dumps
from summarize import summarize
from cpc_analysis import compute_cpc_contribution_ratios
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# 如果本地没有 ace_tools，就定义一个简单的 fallback

# === 日报输出配置 ===
THRESHOLD = 30   # 暴涨/暴跌判定阈值 %
CORE_FIELDS = ["消费金额", "打卡人数", "新增收藏人数", "新好评数", "新中差评数"]
RANK_PLATFORMS = {
    "dianping_hot": "点评热门榜",
    "dianping_checkin": "点评打卡人气榜",
    "dianping_rating":"点评好评榜"
}

try:
    from ace_tools import display_dataframe_to_user
except ImportError:
    def display_dataframe_to_user(title, df):
        print(f"\n=== {title} ===\n")
        print(df.to_string())
        # 如需保存到 CSV，可改为：
        # df.to_csv(f"./daily_report/{title}.csv", encoding="utf-8-sig", index=True)

from matplotlib import font_manager
# 注册项目内的 SimHei.ttf
font_manager.fontManager.addfont("./fonts/SimHei.ttf")
plt.rcParams['font.family'] = 'SimHei'
plt.rcParams['axes.unicode_minus'] = False

# 全局加载门店映射，plot_comparison_table 可以直接用
store_map = pd.read_sql(
    "SELECT 门店ID AS store_id, 美团门店ID, 推广门店 AS brand_name, 运营师 AS operator FROM store_mapping",
    engine
)

def generate_weekly_comparison_table(brand, report_date):
    # 内部直接用全局 engine、store_map
    """
    生成品牌当日 vs 上周同期（如工作日对比同weekday，上周末三天对比）对比表格，
    指标包含：曝光人数, 访问人数, 购买人数, 消费金额,
    新好评数, 新中差评数, 打卡人数, 扫码人数, 新增收藏人数, 点评星级
    """
    metrics = [
        "曝光人数","访问人数","购买人数","消费金额",
        "新好评数","新中差评数","打卡人数","扫码人数","新增收藏人数","点评星级"
    ]
    # 1) MeiTuan IDs for brand
    mids = store_map.loc[store_map['brand_name']==brand, '美团门店ID'].astype(str).tolist()
    if not mids:
        print(f"⚠️ 品牌 {brand} 无对应美团门店ID，跳过")
        return
    mids_sql = ",".join(f"'{m}'" for m in mids)

    # 2) Determine current and last-week dates
    wd = report_date.weekday()  # 0=Mon, ...,6=Sun
    if wd == 0:
        # Monday: compare last Fri-Sun vs Fri-Sun
        curr_start = report_date - timedelta(days=3)
        prev_start = report_date - timedelta(days=10)
        current_dates = [d.date() for d in pd.date_range(curr_start, periods=3)]
        prev_dates    = [d.date() for d in pd.date_range(prev_start, periods=3)]
    else:
        # Tue-Fri: compare that weekday
        current_dates = [report_date.date()]
        prev_dates    = [(report_date - timedelta(days=7)).date()]

    # 3) Load data
    cols = ", ".join(f"`{m}`" for m in metrics)
    cd_sql = ",".join(f"'{d}'" for d in current_dates)
    pd_sql = ",".join(f"'{d}'" for d in prev_dates)
    df_curr = pd.read_sql(
        f"SELECT `日期`, {cols} FROM `operation_data` "
        f"WHERE `美团门店ID` IN ({mids_sql}) AND `日期` IN ({cd_sql}) "
        f"ORDER BY `日期`",
        engine
    ).set_index('日期')
    df_prev = pd.read_sql(
        f"SELECT `日期`, {cols} FROM `operation_data` "
        f"WHERE `美团门店ID` IN ({mids_sql}) AND `日期` IN ({pd_sql}) "
        f"ORDER BY `日期`",
        engine
    ).set_index('日期')
    df_prev = df_prev.add_prefix('上周_')
    df_curr = df_curr.add_prefix('本周_')

    # 4) Combine and compute change rates
    df_cmp = pd.concat([df_curr, df_prev], axis=1)
    for m in metrics:
        curr_col = f"本周_{m}"
        prev_col = f"上周_{m}"
        df_cmp[f"{m}_变化率"] = (
            (df_cmp[curr_col] - df_cmp[prev_col]) / df_cmp[prev_col] * 100
        ).round(1).astype(str) + '%'

    # 5) Display interactive table
    display_dataframe_to_user(f"{brand} 同期对比表", df_cmp)

def plot_vertical_table(brand, report_date, engine):
    from datetime import timedelta

    # 1) 取最近 7 天数据，SQL 里把所有指标都拉出来
    start = report_date - timedelta(days=6)
    df = pd.read_sql(f"""
        SELECT 
            `日期`,
            `消费金额` AS `消费金额`,
            `曝光人数`,
            `访问人数`,
            `购买人数`,
            `打卡人数`,
            `扫码人数`,
            `新增收藏人数`,
            `新好评数`,
            `新中差评数`,
            `点评星级`    AS `星级`
        FROM `operation_data`
        WHERE `美团门店ID` IN (
            SELECT `美团门店ID` 
              FROM `store_mapping` 
             WHERE `推广门店` = '{brand}'
        )
          AND `日期` BETWEEN '{start.date()}' AND '{report_date.date()}'
        ORDER BY `日期`
    """, engine)
    if df.empty:
        print(f"⚠️ {brand} 最近7天无数据，跳过")
        return

    # 2) 把日期列转成纯 YYYY-MM-DD 的字符串，去掉 “00:00:00”
    df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y-%m-%d')

    # 3) 衍生星期几
    weekday_map = {0:'星期一',1:'星期二',2:'星期三',3:'星期四',
                   4:'星期五',5:'星期六',6:'星期日'}
    df['星期'] = pd.to_datetime(df['日期']).dt.weekday.map(weekday_map)

    # 4) 强制把所有“人数”列和评分列变成整数
    int_cols = [
        '曝光人数','访问人数','购买人数',
        '打卡人数','扫码人数','新增收藏人数',
        '新好评数','新中差评数'
    ]
    df[int_cols] = df[int_cols].astype(int)

    # 5) 计算两段转化率
    df['曝光-访问转化率'] = (df['访问人数'] / df['曝光人数'] * 100).round(1).astype(str) + '%'
    df['访问-购买转化率'] = (df['购买人数'] / df['访问人数'] * 100).round(1).astype(str) + '%'

    # 6) 最终选列 & 排序，并重命名为短标题
    cols = [
        '日期', '星期', '消费金额', '曝光人数', '访问人数', '购买人数',
        '曝光-访问转化率', '访问-购买转化率',
        '新增收藏人数', '打卡人数',
        '新好评数', '新中差评数', '扫码人数', '星级'
    ]
    df = df[cols].rename(columns={
        '消费金额': '消费',
        '曝光人数': '曝光',
        '访问人数': '访问',
        '购买人数': '购买',
        '曝光-访问转化率': '访问转化',
        '访问-购买转化率': '购买转化',
        '新增收藏人数': '收藏',
        '打卡人数': '打卡',
        '新好评数': '好评',
        '新中差评数': '差评',
        '扫码人数': '扫码',
        '星级':'星级'
    })

    # 7) 绘制表格，宽度调大，并指定每列相对宽度
    col_widths = []
    for c in df.columns:
        if c in ['日期', '星期']:
            col_widths.append(0.08)
        elif c in ['消费', '曝光', '访问', '购买', '访问转化', '购买转化']:
            col_widths.append(0.06)
        else:
            col_widths.append(0.05)

    fig, ax = plt.subplots(figsize=(16, 0.6 * len(df) + 1))
    ax.axis('off')
    tbl = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        colWidths=col_widths,
        cellLoc='center',
        loc='center'
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(11)
    tbl.scale(1, 1.2)

    ax.set_title(f"{brand} 最近7天关键指标", fontsize=16, pad=12)
    plt.tight_layout()
    out = f"./daily_report/{brand}_最近7天指标表.png"
    fig.savefig(out, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"✅ 保存：{out}")



def plot_comparison_table(brand, report_date):
    """
    生成品牌“本期 vs 上期”同期对比表格并保存成 PNG：
    - 周一对比周五~周日三个工作日，周二~周五对比上周同日
    - 指标：曝光人数, 访问人数, 购买人数, 消费金额,
      新好评数, 新中差评数, 打卡人数, 扫码人数, 新增收藏人数, 点评星级
    """
    # 1) 从全局 store_map 取出本品牌所有美团门店ID
    mids = store_map.loc[store_map['brand_name'] == brand, '美团门店ID'].astype(str).tolist()
    if not mids:
        print(f"⚠️ 品牌 {brand} 无对应美团门店ID，跳过对比表")
        return
    mids_sql = ",".join(f"'{m}'" for m in mids)

    # 2) 计算“本期”“上期”日期
    wd = report_date.weekday()  # 0=周一 … 6=周日
    if wd == 0:
        # 周一：本期 = 周五~周日， 上期 = 上周周五~周日
        curr_start = report_date - timedelta(days=3)
        prev_start = report_date - timedelta(days=10)
        current_dates = [d.date() for d in pd.date_range(curr_start, periods=3)]
        prev_dates    = [d.date() for d in pd.date_range(prev_start, periods=3)]
        label_curr, label_prev = "本期(周五~周日)", "上期(上周周五~周日)"
    else:
        # 周二~周五：本期 = 昨日, 上期 = 上周同期
        current_dates = [(report_date - timedelta(days=1)).date()]
        prev_dates    = [(report_date - timedelta(days=8)).date()]
        label_curr, label_prev = "昨日", "上周同期"

    # 3) 拉取数据
    metrics = [
        "曝光人数","访问人数","购买人数","消费金额",
        "新好评数","新中差评数","打卡人数","扫码人数","新增收藏人数","点评星级"
    ]
    cols = ", ".join(f"`{m}`" for m in metrics)
    cd_sql = ", ".join(f"'{d}'" for d in current_dates)
    pd_sql = ", ".join(f"'{d}'" for d in prev_dates)

    df_curr = pd.read_sql(
        f"SELECT `日期`, {cols} FROM `operation_data` "
        f"WHERE `美团门店ID` IN ({mids_sql}) AND `日期` IN ({cd_sql})",
        engine
    ).set_index("日期")

    df_prev = pd.read_sql(
        f"SELECT `日期`, {cols} FROM `operation_data` "
        f"WHERE `美团门店ID` IN ({mids_sql}) AND `日期` IN ({pd_sql})",
        engine
    ).set_index("日期")

    # 4) 汇总并构造对比表
    if df_curr.empty or df_prev.empty:
        print(f"⚠️ 品牌 {brand} 本期或上期无数据，跳过对比表")
        return
    curr_sum = df_curr.sum()
    prev_sum = df_prev.sum()
    df_cmp = pd.DataFrame({
        "指标": metrics,
        label_curr: [curr_sum[m] for m in metrics],
        label_prev: [prev_sum[m] for m in metrics],
    })
    df_cmp["变化率"] = (
        (df_cmp[label_curr] - df_cmp[label_prev]) / df_cmp[label_prev] * 100
    ).round(1).astype(str) + "%"

    # 5) 展示并保存
    display_dataframe_to_user(f"{brand} 同期对比表", df_cmp)

    fig, ax = plt.subplots(figsize=(len(df_cmp)*0.8, 2.5))
    ax.axis("off")
    tbl = ax.table(
        cellText=df_cmp.values,
        colLabels=df_cmp.columns,
        loc="center"
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(12)
    tbl.scale(1, 1.5)
    ax.set_title(brand, fontsize=16, pad=10)
    plt.tight_layout()
    out_path = f"./daily_report/{brand}_同期对比表.png"
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"✅ 保存 {brand} 同期对比表：{out_path}")

def main():

    # ---------- CLI & 日期计算 ----------
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="日报日期，默认昨天 (YYYY-MM-DD)")
    args = parser.parse_args()

    # 1) 先计算默认拉取的日期（昨天）
    default_dt = datetime.now() - timedelta(days=1)
    default_str = default_dt.strftime("%Y-%m-%d")

    # 2) 如果命令行给了 --date，就直接用；否则弹交互提示
    if args.date:
        date_str = args.date
    else:
        # 提示用户确认或输入
        inp = input(f"使用日期 [{default_str}]？回车确认 或 输入其他日期 (YYYY-MM-DD)：").strip()
        date_str = inp if inp else default_str

    # 3) 最后解析
    try:
        report_date = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        print(f"❌ 日期格式不对：{date_str}，请按 YYYY-MM-DD 重试")
        return

    last_7_start = report_date - timedelta(days=7)
    print(f"🗓️ 最终使用日期：{report_date.date()}，生成日报")

    # ---------- 在开始写 TXT/Excel 之前，确保输出目录存在 ----------
    out_dir = Path("./daily_report")
    out_dir.mkdir(exist_ok=True)

    # ---------- 一次性拉取所有数据 ----------
    print("➡️ 拉取昨日运营数据")
    op_today = pd.read_sql(
        f"""SELECT 
               `美团门店ID`,`曝光人数`,`访问人数`,`购买人数`,
               `消费金额`,`新好评数`,`新中差评数`,
               `打卡人数`,`扫码人数`,`点评星级`,`新增收藏人数`,`rankings_detail`
           FROM `operation_data`
           WHERE `日期` = '{report_date.date()}'""",
        engine
    ).merge(
        store_map[['美团门店ID','brand_name','operator']],
        on='美团门店ID', how='left'
    ).rename(columns={'brand_name':'推广门店'})

    print("➡️ 拉取近7天运营数据")
    op_last7 = pd.read_sql(
        f"""SELECT
               `美团门店ID`,`日期`,`曝光人数`,`访问人数`,`购买人数`,
               `消费金额`,`新好评数`,`新中差评数`,
               `打卡人数`,`扫码人数`,`点评星级`,`新增收藏人数`
           FROM `operation_data`
           WHERE `日期` BETWEEN '{last_7_start.date()}' AND '{report_date.date()}'""",
        engine
    ).merge(
        store_map[['美团门店ID','brand_name','operator']],
        on='美团门店ID', how='left'
    ).rename(columns={'brand_name':'推广门店'})

    print("➡️ 拉取昨日CPC数据")
    cpc_today = pd.read_sql(
        f"""SELECT `store_id`,`cost`,`impressions`,`clicks`,`orders`
           FROM `cpc_hourly_data`
           WHERE `date` = '{report_date.date()}'""",
        engine
    ).merge(
        store_map[['store_id','brand_name','operator']],
        on='store_id', how='left'
    ).rename(columns={'brand_name':'推广门店'})

    # ---------- 加载最近 14 天历史数据（用于环比 & 表格） ----------
    print("➡️ 拉取最近14天运营数据")
    op_hist = pd.read_sql(
        f"""SELECT 
               `美团门店ID`,`日期`,`曝光人数`,`访问人数`,`购买人数`,
               `消费金额`,`新好评数`,`新中差评数`,
               `打卡人数`,`扫码人数`,`点评星级`,`新增收藏人数`
           FROM `operation_data`
           WHERE `日期` BETWEEN '{(report_date - timedelta(days=14)).date()}' AND '{report_date.date()}'""",
        engine
    ).merge(
        store_map[['美团门店ID','brand_name','operator']],
        on='美团门店ID', how='left'
    ).rename(columns={'brand_name':'推广门店'})

    # ---------- 分组准备 ----------
    op_group  = op_today.groupby("推广门店")
    op7_group = op_last7.groupby("推广门店")
    cpc_group = cpc_today.groupby("推广门店")

    # ---------- 先为每家门店生成对比图表 ----------
    #for brand in op_group.groups.keys():
        #plot_vertical_table(brand, report_date, engine)

    # ---------- 再为每家门店生成同期对比表 ----------
    #for brand in op_group.groups.keys():
        #generate_weekly_comparison_table(brand, report_date)

    # ---------- 通用格式化 ----------
    def is_big_change(pct_str, threshold=THRESHOLD):
        try:
            return abs(float(pct_str.strip('%'))) >= threshold
        except:
            return False

    def fmt(val, pct_str):
        if pct_str == "N/A":
            return str(val)
        pct = float(pct_str.strip('%'))
        if pct > 0:
            return f"{val}(+{pct}%)"
        if pct < 0:
            return f"{val}(-{abs(pct)}%)"
        return f"{val}(持平)"

    from collections import defaultdict

    operator_sections = defaultdict(list)

    # ---------- 构造每家门店的 Section ----------
    sections = []
    for brand, df_op in op_group:
        # 准备当日7天历史 & CPC & 全量历史数据
        df_op7    = op7_group.get_group(brand) if brand in op7_group.groups else pd.DataFrame()
        df_cpc    = cpc_group.get_group(brand)  if brand in cpc_group.groups  else pd.DataFrame()
        df_hist_b = op_hist[op_hist["推广门店"] == brand]

        # 指标字段
        op_fields  = ["曝光人数","访问人数","购买人数","消费金额",
                      "新好评数","新中差评数","打卡人数","扫码人数","点评星级","新增收藏人数"]
        cpc_fields = ["cost","impressions","clicks","orders"]

        # 1) 汇总当日运营 & CPC 数据
        op_sum = summarize(df_op, op_fields)
        if df_cpc.empty:
            cpc_sum, ratios = {}, {}
        else:
            cpc_sum = summarize(df_cpc, cpc_fields)
            ratios  = compute_cpc_contribution_ratios(op_sum, cpc_sum)

        # 2) 计算日环比（昨日 vs. 上周同期）
        weekday = report_date.weekday()
        if weekday == 0:
            curr = df_hist_b[
                (df_hist_b["日期"] >= (report_date - timedelta(days=3)).date()) &
                (df_hist_b["日期"] <= (report_date - timedelta(days=1)).date())
            ]
            prev = df_hist_b[
                (df_hist_b["日期"] >= (report_date - timedelta(days=10)).date()) &
                (df_hist_b["日期"] <= (report_date - timedelta(days=8)).date())
            ]
        else:
            curr = df_hist_b[df_hist_b["日期"] == (report_date - timedelta(days=1)).date()]
            prev = df_hist_b[df_hist_b["日期"] == (report_date - timedelta(days=8)).date()]

        curr_sum = summarize(curr, op_fields)
        prev_sum = summarize(prev, op_fields)
        link_ratio = {}
        for k in op_fields:
            if prev_sum.get(k,0):
                ratio = round(
                    (curr_sum.get(k,0) - prev_sum.get(k,0))
                    / (prev_sum.get(k,0) or 1)
                    * 100, 1
                )
                link_ratio[k] = f"{ratio}%"
            else:
                link_ratio[k] = "N/A"

        # 3) 计算与 7 天均值环比（可留存但不在最终输出中）
        op7_sum = summarize(df_op7, op_fields)
        cmp7 = {
            k: f"{round((op_sum.get(k,0) - op7_sum.get(k,0)) /
                        (op7_sum.get(k,1) or 1) * 100, 1)}%"
            if op7_sum.get(k) else "N/A"
            for k in op_fields
        }

        # 4) 解析榜单动态
        non_null = df_op["rankings_detail"].dropna().tolist() if not df_op.empty else []
        raw_rank = non_null[0] if non_null else None
        if isinstance(raw_rank, (bytes, bytearray)):
            try: raw_rank = raw_rank.decode("utf-8")
            except: raw_rank = None
        if isinstance(raw_rank, str):
            try: parsed = json.loads(raw_rank)
            except: parsed = {}
        elif isinstance(raw_rank, dict):
            parsed = raw_rank
        else:
            parsed = {}
        if isinstance(parsed, str):
            try: tmp = json.loads(parsed)
            except: tmp = {}
            parsed = tmp if isinstance(tmp, dict) else {}
        rank_dict = parsed if isinstance(parsed, dict) else {}

        # 只保留这几个平台
        allowed = ["dianping_hot", "dianping_checkin", "dianping_rating"]
        # 展示阈值
        th_city = 10
        th_sub = 5
        level_map = {"city": "全市榜", "subdistrict": "区县榜", "business": "商圈榜"}

        rank_lines = []
        for pf in allowed:
            scope = rank_dict.get(pf, {})
            if not isinstance(scope, dict):
                continue
            desc = RANK_PLATFORMS[pf]
            # 判断各级是否达标
            city_rank = scope.get("city")
            sub_rank = scope.get("subdistrict")
            bus_rank = scope.get("business")
            # 是否展示 city / subdistrict
            if isinstance(city_rank, int) and city_rank <= th_city:
                rank_lines.append(f"{desc}{level_map['city']}第{city_rank}名")
            if isinstance(sub_rank, int) and sub_rank <= th_sub:
                rank_lines.append(f"{desc}{level_map['subdistrict']}第{sub_rank}名")
            # 如果 city 和 subdistrict 都不达标，只展示商圈榜
            if not (
                    (isinstance(city_rank, int) and city_rank <= th_city) or
                    (isinstance(sub_rank, int) and sub_rank <= th_sub)
            ) and isinstance(bus_rank, int):
                rank_lines.append(f"{desc}{level_map['business']}第{bus_rank}名")

        if rank_lines:
            rank_note = "\n".join(rank_lines)
        else:
            rank_note = "无榜单变化"

        # 5) 取昨日四项核心指标 & 环比
        today    = df_op.iloc[0]
        card_val = int(today['打卡人数'])
        good_val = int(today['新好评数'])
        bad_val  = int(today['新中差评数'])
        rev_val  = round(today['消费金额'], 0)

        card_pct = link_ratio.get('打卡人数',       'N/A')
        good_pct = link_ratio.get('新好评数',       'N/A')
        bad_pct  = link_ratio.get('新中差评数',     'N/A')
        rev_pct  = link_ratio.get('消费金额', 'N/A')


        # 用 fmt() 格式化，并处理“差评为0”场景
        card_str = fmt(card_val, card_pct)
        good_str = fmt(good_val, good_pct)
        bad_str  = "暂无差评" if bad_val == 0 else fmt(bad_val, bad_pct)
        rev_str  = fmt(rev_val, rev_pct)

        # 新增收藏
        col_val = int(op_sum.get("新增收藏人数", 0))
        col_pct = link_ratio.get("新增收藏人数", "N/A")
        col_str = fmt(col_val, col_pct)

        # 异动检测（曝光/访问 ±30% 及以上才提示）
        extra_notes = []
        exp_pct = link_ratio.get("曝光人数", "N/A")
        if is_big_change(exp_pct):
            extra_notes.append(f"⚠️ 曝光{exp_pct}")
        vis_pct = link_ratio.get("访问人数", "N/A")
        if is_big_change(vis_pct):
            extra_notes.append(f"⚠️ 访问{vis_pct}")
        extra_note_str = "；" + "；".join(extra_notes) if extra_notes else ""

        # 6) 拼装两行输出
        suggestion = ("发现差评，运营师已介入跟进。"
                      if bad_val > 0
                      else "数据正常，继续引导好评。")
        # 新增收藏
        col_val = int(op_sum.get("新增收藏人数", 0))
        col_pct = link_ratio.get("新增收藏人数", "N/A")
        col_str = fmt(col_val, col_pct)

        # 异动提示
        notes = []
        for k in ["曝光人数", "访问人数"]:
            pct = link_ratio.get(k, "N/A")
            if is_big_change(pct):
                alias = "曝光" if k == "曝光人数" else "访问"
                notes.append(f"{alias}{pct}")
        note_str = f" ⚠️ {'/'.join(notes)}" if notes else ""

        # 差评
        bad_val = int(op_sum.get("新中差评数", 0))
        bad_str = fmt(bad_val, link_ratio.get("新中差评数", "N/A"))

        # 定义TEXT
        text = (
            f"{brand}\n"
            f"- 核心指标：消费 {rev_str}；打卡 {card_str}；收藏 {col_str}；\n"
            f"  好评 {good_str}；{'' if bad_val == 0 else '差评 ' + bad_str}\n"
            f"- 排行榜：\n{rank_note}\n"
            f"- 建议：{suggestion}\n"
        )

        # 保存到单店 txt（文件名：品牌_日期.txt）
        #out_dir = Path("./daily_report")
        #out_dir.mkdir(exist_ok=True)
        #fn = out_dir / f"{brand}_{report_date.date()}.txt"
        #fn.write_text(text, encoding="utf-8-sig")
        #print(f"📄 已生成：{fn}")

        # 按 operator 收集
        op = store_map.loc[store_map['brand_name'] == brand, 'operator'].iat[0]
        operator_sections[op].append(text)

        sections.append(
            f"{brand}\n"
            f"- 核心指标：消费 {rev_str}；打卡 {card_str}；收藏 {col_str}；\n"
            f"  好评 {good_str}；{'' if bad_val == 0 else '差评 ' + bad_str}\n"
            f"- 排行榜：\n{rank_note}\n"
            f"- 建议：{'👏 好评稳增，继续引导五星' if (good_val > 0 and bad_val == 0) else '差评已转运营师跟进' if bad_val > 0 else '数据正常'}\n"
        )

    # —— 循环外，写各运营师的 TXT ——
    for op, texts in operator_sections.items():
        fn = out_dir / f"{op}_{report_date.date()}.txt"
        fn.write_text("\n\n".join(texts), encoding="utf-8-sig")
        print(f"📄 已生成运营师日报：{fn}")

    # —— 一次性拉当月全量数据 & 衍生列（循环外） ——
    month_start = report_date.replace(day=1).date()
    df_month = pd.read_sql(
        f"""
        SELECT
          `美团门店ID`,`日期`,`消费金额`,`曝光人数`,`访问人数`,`购买人数`,
          `扫码人数`,`新增收藏人数`,`打卡人数`,`新好评数`,`新中差评数`,`点评星级`
        FROM `operation_data`
        WHERE `日期` BETWEEN '{month_start}' AND '{report_date.date()}'
        """, engine
    ).merge(
        store_map[['美团门店ID','brand_name','operator']],
        on='美团门店ID', how='left'
    )
    # 衍生“星期”、“访问转化”、“购买转化”
    weekday_map = {0:'星期一',1:'星期二',2:'星期三',3:'星期四',4:'星期五',5:'星期六',6:'星期日'}
    df_month['星期'] = pd.to_datetime(df_month['日期']).dt.weekday.map(weekday_map)
    df_month['访问转化'] = (df_month['访问人数']/df_month['曝光人数']).round(3)
    df_month['购买转化'] = (df_month['购买人数']/df_month['访问人数']).round(3)
    cols = ['日期','星期','消费金额','曝光人数','访问人数','购买人数',
            '访问转化','购买转化','新增收藏人数','打卡人数',
            '新好评数','新中差评数','扫码人数','点评星级']

    # —— 按运营师输出月度 Excel（每店一个 sheet） ——
    for op, grp_op in df_month.groupby('operator'):
        file = out_dir / f"{op}_{report_date.date()}_月度数据.xlsx"
        # 1) 写每个门店分别一个 sheet
        with pd.ExcelWriter(file, engine='openpyxl', mode='w') as writer:
            for brand, grp in grp_op.groupby('brand_name'):
                sheet = grp.sort_values('日期')[cols]
                sheet.to_excel(
                    writer,
                    sheet_name=brand[:31],  # sheet 名称截 31 字
                    index=False,
                    startrow=2
                )
        print(f"✅ 已生成运营师月度 Excel：{file}")

        # 4) 样式 + 冻结 + 格式化 —— 这里用 file 而非 out_excel
        wb = openpyxl.load_workbook(file)
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        for ws in wb.worksheets:
            max_col = ws.max_column

            # ——— 合并第1行写门店名称 ———
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            title = ws.cell(row=1, column=1)
            title.value = ws.title
            title.font = Font(size=14, bold=True)
            title.alignment = Alignment(horizontal="center", vertical="center")

            # ——— 第2行留白 ———
            # （什么都不用写，自动保留空行）

            # ——— 第3行样式：加粗居中＋灰底 ———
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # ——— 冻结前三行 ———
            ws.freeze_panes = "A4"

            # ——— 列宽自适应 & 强制最小宽度 ———
            # 为了让 A 列（日期）不要过宽，D~L 列足够宽，定义最小宽度：
            min_widths = {
                'A': 13,  # 日期
                'B': 8,  # 星期
                'C': 10,  # 消费金额
                'D': 10, 'E': 10, 'F': 10, 'G': 10,
                'H': 10, 'I': 13, 'J': 10, 'K': 10, 'L': 11,
                'M': 8,  # 扫码
                'N': 8  # 星级
            }
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                # 如果是日期列，直接用固定宽度
                if col_letter == 'A':
                    ws.column_dimensions[col_letter].width = min_widths['A']
                    continue

                # 其它列照常自适应，但保证不小于 min_widths
                max_len = max(
                    len(str(c.value)) if c.value is not None else 0
                    for c in col_cells
                )
                optimal = max_len + 4
                ws.column_dimensions[col_letter].width = max(
                    optimal, min_widths.get(col_letter, optimal)
                )

            # ——— 单元格格式化 ———
            money_fmt = '#,##0'
            pct_fmt = '0.0%'
            star_fmt = '0.0'
            # “点评星级”在最后一列，这里用 max_column 保证定位
            for row in range(4, ws.max_row + 1):
                ws.cell(row, 3).number_format = money_fmt  # C 列：消费金额
                ws.cell(row, 7).number_format = pct_fmt  # G 列：访问转化
                ws.cell(row, 8).number_format = pct_fmt  # H 列：购买转化
                ws.cell(row, ws.max_column).number_format = star_fmt  # N 列：星级

            # ——— 单元格格式化 ———
            money_fmt = '#,##0'
            pct_fmt = '0.0%'
            star_fmt = '0.0'
            # 假设列序：C=消费金额, G=访问转化, H=购买转化, N=点评星级
            ws.column_dimensions['C'].width += 2
            for row in range(4, ws.max_row + 1):
                ws.cell(row, 3).number_format = money_fmt
                ws.cell(row, 7).number_format = pct_fmt
                ws.cell(row, 8).number_format = pct_fmt
                ws.cell(row, ws.max_column).number_format = star_fmt

        wb.save(file)
        print(f"✅ Excel 已格式化：{file}")

    # ---------- 每 5 家一组调用 AI 输出 ----------
    '''md_chunks = []
    group_size = 5
    header = (
        "你是餐饮BI运营助手，请基于以下品牌昨日关键数据，"
        "用 Markdown 输出，每家店仅两行：\n"
        "第一行：昨日打卡人数、好评数、中差评数、消费金额环比；\n"
        "第二行：给出榜单信息，并附上“请店内同事……”式的行动建议；\n"
        "其余不必赘述，全部用中文。\n\n"
    )

    for idx in range(0, len(sections), group_size):
        batch    = sections[idx:idx+group_size]
        prompt   = header + "\n\n".join(batch)
        batch_no = idx // group_size + 1
        print(f"➡️ 调用 AI（第{batch_no}批）")
        for attempt in range(1, 4):
            try:
                md = call_kimi_api([{"role":"user","content":prompt}], API_KEY, MODEL, temperature=0.3)
                md_chunks.append(md)
                break
            except SSLError:
                print(f"⚠️ AI 调用第{batch_no}批第{attempt}次失败，重试…")
                time.sleep(attempt * 2)
        else:
            print(f"❌ 第{batch_no}批最终失败，跳过该批内容")
            md_chunks.append(f"### 第{batch_no}批内容生成失败，请稍后查看。")

    # 合并并写文件
    md_all = "\n\n---\n\n".join(md_chunks)
    out_dir = Path("./daily_report")
    out_dir.mkdir(exist_ok=True)
    fp = out_dir / f"{report_date.date()}.md"
    fp.write_text(md_all, encoding="utf-8")
    print("✅ 日报生成：", fp)'''

if __name__ == "__main__":
    main()
